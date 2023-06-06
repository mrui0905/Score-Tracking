import openpyxl

wb = openpyxl.load_workbook("Zetamac.xlsx")

add = wb['Addition']
subtract = wb['Subtraction']
multiply = wb['Multiplication']
divide = wb['Division']
combined = wb['Combined']


curr = 2

def prompt():
    global curr

    print("Select 0 for addition, 1 for subtraction, 2 for multiplcation, 3 for division, 4 for combined")
    mode = int(input("Mode: "))

    curr = 2
    while [add, subtract, multiply, divide, combined][mode].cell(row = curr, column = 2).value:
        curr += 1

    return [add, subtract, multiply, divide, combined][mode]

def length():
    return input("Length: ")

def time():
    return 'AM' if input("Session: ") == 'AM' else 'PM'

def averages():
    global session
    global date

    for name in wb.sheetnames:
        sheet = wb[name]

        curr = 2
        while sheet.cell(row = curr, column = 2).value:
            curr += 1

        if date == sheet.cell(row = curr-1, column = 2).value and session == sheet.cell(row = curr-1, column = 3).value:
            update_averages(sheet, curr-1)
        

def update_averages(sheet, curr):
    global session
    global date
    tot = 0
    count = 0

    while sheet.cell(row = curr, column = 2).value == date and session == sheet.cell(row = curr, column = 3).value:
        tot += int(sheet.cell(row = curr, column = 4).value)
        count += int(sheet.cell(row = curr, column = 5).value)
        curr += -1

    average_row = 2
    while sheet.cell(row = average_row, column = 8).value:
        average_row += 1
    
    sheet.cell(row = average_row, column = 8).value = date
    sheet.cell(row = average_row, column = 9).value = round(tot/(count*60), 4)

# prompt for date, sheet, and length
date = input("Day: ")
sheet = prompt()
len = length()
session = time()






print("Input 'end' to end program, 'len' to change time length, 'mode' to change game mode, 'session' to change session, or your score to record your score.")

while True:
    text = input("Input: ")

    if text == 'end':
        averages()
        wb.save("Zetamac.xlsx")
        print("End Program")
        break

    if text == 'len':
        len = length()
        continue

    if text == 'mode':
        sheet = prompt()
        continue

    if text == 'session':
        session = time()
        continue

    if not text.isdigit():
        continue

    sheet.cell(row = curr, column = 2).value = date
    sheet.cell(row = curr, column = 3).value = session
    sheet.cell(row = curr, column = 4).value = text
    sheet.cell(row = curr, column = 5).value = len
    curr+= 1








    








